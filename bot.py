import os
import io
import math
import shutil
import calendar
import smtplib
from datetime import datetime, date, timedelta, time as dt_time
from itertools import groupby
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import pytz
from dotenv import load_dotenv
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    CallbackContext,
)

from database import Database

load_dotenv()

TOKEN = os.getenv("TELEGRAM_TOKEN")
ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip()]
TIMEZONE = pytz.timezone("America/Argentina/Buenos_Aires")

db = Database()

# Verificación por geolocalización (opcional)
REQUIRE_LOCATION  = os.getenv("REQUIRE_LOCATION", "false").lower() == "true"
OFFICE_LAT        = float(os.getenv("OFFICE_LAT", "0") or "0")
OFFICE_LON        = float(os.getenv("OFFICE_LON", "0") or "0")
OFFICE_RADIUS_M   = float(os.getenv("OFFICE_RADIUS_METERS", "300") or "300")

PALABRAS_ENTRADA = [
    "entro", "entré", "entre", "llegué", "llegue",
    "buenos días", "buenos dias", "buen dia", "buen día",
    "inicio", "empiezo",
]
PALABRAS_SALIDA = [
    "salgo", "me voy", "hasta mañana", "hasta manana",
    "chau", "me retiro", "salida", "termino", "terminé",
]

AWAITING_NAME = "awaiting_name"

DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

ABSENCE_TYPES = {
    "vacacion": "Vacación", "vacación": "Vacación",
    "enfermedad": "Enfermedad",
    "licencia": "Licencia",
    "justificada": "Justificada",
    "injustificada": "Injustificada",
}


def ahora() -> datetime:
    return datetime.now(TIMEZONE)


def es_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def _distancia_metros(lat1, lon1, lat2, lon2) -> float:
    R = 6_371_000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


async def _pedir_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Pide el nombre al usuario y activa el flag de espera."""
    context.user_data[AWAITING_NAME] = True
    await update.message.reply_text(
        "Hola! Para registrarte necesito tu nombre completo.\n"
        "¿Cómo te llamás?"
    )


async def _guardar_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """Si hay un nombre pendiente, lo guarda y devuelve True."""
    if not context.user_data.get(AWAITING_NAME):
        return False
    name = update.message.text.strip()
    db.register_employee(update.effective_user.id, name)
    context.user_data[AWAITING_NAME] = False
    await update.message.reply_text(
        f"*Bienvenido/a, {name}!*\n"
        f"Ya estás registrado en el sistema de asistencia.\n"
        f"\n"
        f"*Cómo usarlo:*\n"
        f"\n"
        f"Al llegar, escribí cualquiera de estas frases:\n"
        f"  › llegué\n"
        f"  › entro\n"
        f"  › buenos días\n"
        f"  › /entro\n"
        f"\n"
        f"Al irte, escribí cualquiera de estas:\n"
        f"  › me voy\n"
        f"  › salgo\n"
        f"  › chau\n"
        f"  › /salgo\n"
        f"\n"
        f"Para ver cómo vas hoy:\n"
        f"  › /estado\n"
        f"\n"
        f"_Eso es todo. Sin formularios, sin papel._\n"
        f"_Cualquier duda hablá con tu administrador._",
        parse_mode="Markdown",
    )
    return True


# ---------- handlers ----------

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    employee = db.get_employee(update.effective_user.id)
    if employee:
        await update.message.reply_text(
            f"Ya estás registrado como *{employee['name']}*\n\n"
            "  /entro  - Registrar entrada\n"
            "  /salgo  - Registrar salida\n"
            "  /estado - Ver tu estado hoy",
            parse_mode="Markdown",
        )
        return
    await _pedir_nombre(update, context)


async def _hacer_entro(update: Update, context: ContextTypes.DEFAULT_TYPE = None,
                       ubicacion_verificada: bool = False):
    user = update.effective_user
    employee = db.get_employee(user.id)
    if not employee:
        if context:
            await _pedir_nombre(update, context)
        else:
            await update.message.reply_text("Primero registrate enviando /start")
        return
    if REQUIRE_LOCATION and not ubicacion_verificada:
        await update.message.reply_text(
            "Para registrar entrada compartí tu ubicación.\n"
            "Tocá el clip > Ubicación > Compartir ubicación."
        )
        return

    ts = ahora()
    result = db.register_entry(user.id, ts)

    if result == "already_in":
        status = db.get_today_status(user.id, ts.date())
        hora = status["entry_time"].strftime("%H:%M") if status else "?"
        await update.message.reply_text(
            f"Ya registraste entrada hoy a las {hora}."
        )
        return

    await update.message.reply_text(
        f"*Entrada registrada*\n"
        f"Nombre: {employee['name']}\n"
        f"Hora:   {ts.strftime('%H:%M')}\n"
        f"Fecha:  {ts.strftime('%d/%m/%Y')}",
        parse_mode="Markdown",
    )


async def _hacer_salgo(update: Update, context: ContextTypes.DEFAULT_TYPE = None,
                       ubicacion_verificada: bool = False):
    user = update.effective_user
    employee = db.get_employee(user.id)
    if not employee:
        if context:
            await _pedir_nombre(update, context)
        else:
            await update.message.reply_text("Primero registrate enviando /start")
        return
    if REQUIRE_LOCATION and not ubicacion_verificada:
        await update.message.reply_text(
            "Para registrar salida compartí tu ubicación.\n"
            "Tocá el clip > Ubicación > Compartir ubicación."
        )
        return

    ts = ahora()
    result = db.register_exit(user.id, ts)

    if result is None:
        await update.message.reply_text(
            "No tenés entrada registrada hoy. Usá /entro primero."
        )
        return

    horas = result["total_hours"]
    await update.message.reply_text(
        f"*Salida registrada*\n"
        f"Nombre:  {employee['name']}\n"
        f"Hora:    {ts.strftime('%H:%M')}\n"
        f"Trabajaste {horas:.1f} hs hoy.",
        parse_mode="Markdown",
    )


async def cmd_entro(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await _hacer_entro(update, context)


async def cmd_salgo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await _hacer_salgo(update, context)


async def cmd_estado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    employee = db.get_employee(user.id)
    if not employee:
        await _pedir_nombre(update, context)
        return

    ts = ahora()
    status = db.get_today_status(user.id, ts.date())

    if not status:
        msg = f"*{employee['name']}*\nHoy todavía no registraste entrada."
    elif status["exit_time"] is None:
        msg = (
            f"*{employee['name']}*\n"
            f"Dentro desde las {status['entry_time'].strftime('%H:%M')}"
        )
    else:
        msg = (
            f"*{employee['name']}*\n"
            f"Jornada cerrada - {status['total_hours']:.1f} hs trabajadas\n"
            f"Entrada: {status['entry_time'].strftime('%H:%M')} | "
            f"Salida: {status['exit_time'].strftime('%H:%M')}"
        )

    await update.message.reply_text(msg, parse_mode="Markdown")


async def cmd_reporte(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if not es_admin(user.id):
        await update.message.reply_text(
            "Solo los administradores pueden generar reportes."
        )
        return

    hoy    = ahora().date()
    start  = date(hoy.year, 1, 1)            # 1 de enero
    end    = date(hoy.year, 12, 31)          # 31 de diciembre
    titulo = f"Reporte Anual {hoy.year}"

    records  = db.get_records_by_period(start, end)
    holidays = db.get_holidays(start, end)
    absences = db.get_absences(start, end)
    buffer   = _build_xlsx(records, titulo, start_date=start, end_date=end,
                           holidays=holidays, absences=absences)
    filename = f"asistencia_{hoy.year}.xlsx"
    await update.message.reply_document(
        document=buffer,
        filename=filename,
        caption=f"Reporte anual {hoy.year} — todos los días del año",
    )

    # También por email si está configurado
    buffer.seek(0)
    email_to = os.getenv("EMAIL_TO", "")
    if _enviar_email(buffer, filename, f"Asistencia - Reporte {ahora().strftime('%d/%m/%Y')}"):
        await update.message.reply_text(f"Reporte también enviado a {email_to}")


async def cmd_empleados(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if not es_admin(user.id):
        await update.message.reply_text(
            "Solo los administradores pueden ver la lista de empleados."
        )
        return

    empleados = db.list_employees()
    if not empleados:
        await update.message.reply_text("No hay empleados registrados aún.")
        return

    lines = [f"*Empleados registrados ({len(empleados)}):*"]
    for e in empleados:
        lines.append(f"  - {e['name']} (ID: {e['telegram_id']})")

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def handle_location(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Procesa ubicación compartida para fichaje con verificación geográfica."""
    user = update.effective_user
    if await _guardar_nombre(update, context):
        return
    employee = db.get_employee(user.id)
    if not employee:
        await _pedir_nombre(update, context)
        return

    if not OFFICE_LAT and not OFFICE_LON:
        await update.message.reply_text(
            "La verificación por ubicación no está configurada.\n"
            "Usá /entro o /salgo directamente."
        )
        return

    loc  = update.message.location
    dist = _distancia_metros(loc.latitude, loc.longitude, OFFICE_LAT, OFFICE_LON)

    if dist > OFFICE_RADIUS_M:
        await update.message.reply_text(
            f"Estás a {dist:.0f} m de la oficina "
            f"(límite: {OFFICE_RADIUS_M:.0f} m). No se puede registrar."
        )
        return

    ts     = ahora()
    status = db.get_today_status(user.id, ts.date())
    if not status or status["exit_time"] is not None:
        await _hacer_entro(update, context, ubicacion_verificada=True)
    else:
        await _hacer_salgo(update, context, ubicacion_verificada=True)


# ---------- comandos admin: corrección de registros ----------

async def cmd_corregir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Uso: /corregir Juan Perez entrada 08:30 25/06
            /corregir Juan Perez salida  17:00 25/06"""
    if not es_admin(update.effective_user.id):
        return
    # args: [...nombre..., entrada|salida, HH:MM, DD/MM]
    if len(context.args) < 4:
        await update.message.reply_text(
            "Uso: /corregir Nombre entrada HH:MM DD/MM\n"
            "     /corregir Nombre salida  HH:MM DD/MM"
        )
        return

    fecha_raw  = context.args[-1]
    hora_raw   = context.args[-2]
    tipo       = context.args[-3].lower()
    nombre     = " ".join(context.args[:-3])

    if tipo not in ("entrada", "salida"):
        await update.message.reply_text("El tipo debe ser 'entrada' o 'salida'.")
        return

    try:
        day, month  = fecha_raw.split("/")
        h, m        = hora_raw.split(":")
        d           = date(ahora().year, int(month), int(day))
        nueva_hora  = TIMEZONE.localize(
            datetime(d.year, d.month, d.day, int(h), int(m))
        )
    except Exception:
        await update.message.reply_text("Formato inválido. Ejemplo: /corregir Juan entrada 08:30 25/06")
        return

    matches = db.find_employee_by_name(nombre)
    if not matches:
        await update.message.reply_text("No encontré ese empleado.")
        return
    if len(matches) > 1:
        await update.message.reply_text(
            f"Varios resultados: {', '.join(m['name'] for m in matches)}. Sé más específico."
        )
        return

    emp   = matches[0]
    field = "entry_time" if tipo == "entrada" else "exit_time"
    ok    = db.update_attendance_field(emp["telegram_id"], d, field, nueva_hora)

    if ok:
        await update.message.reply_text(
            f"Corregido: {emp['name']} — {tipo} {hora_raw} del {d.strftime('%d/%m/%Y')}"
        )
    else:
        await update.message.reply_text(
            f"No hay registro de {emp['name']} el {d.strftime('%d/%m/%Y')}."
        )


async def cmd_borrar_fichaje(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Uso: /borrar Juan Perez 25/06"""
    if not es_admin(update.effective_user.id):
        return
    if len(context.args) < 2:
        await update.message.reply_text("Uso: /borrar Nombre DD/MM")
        return

    fecha_raw = context.args[-1]
    nombre    = " ".join(context.args[:-1])

    try:
        day, month = fecha_raw.split("/")
        d = date(ahora().year, int(month), int(day))
    except Exception:
        await update.message.reply_text("Formato inválido. Ejemplo: /borrar Juan 25/06")
        return

    matches = db.find_employee_by_name(nombre)
    if not matches:
        await update.message.reply_text("No encontré ese empleado.")
        return
    if len(matches) > 1:
        await update.message.reply_text(
            f"Varios resultados: {', '.join(m['name'] for m in matches)}. Sé más específico."
        )
        return

    emp = matches[0]
    ok  = db.delete_attendance(emp["telegram_id"], d)
    if ok:
        await update.message.reply_text(
            f"Registro eliminado: {emp['name']} — {d.strftime('%d/%m/%Y')}"
        )
    else:
        await update.message.reply_text(
            f"No había registro de {emp['name']} el {d.strftime('%d/%m/%Y')}."
        )


async def cmd_ver_fichajes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Uso: /ver Juan Perez — muestra los últimos 10 registros."""
    if not es_admin(update.effective_user.id):
        return
    if not context.args:
        await update.message.reply_text("Uso: /ver Nombre")
        return

    matches = db.find_employee_by_name(" ".join(context.args))
    if not matches:
        await update.message.reply_text("No encontré ese empleado.")
        return
    if len(matches) > 1:
        await update.message.reply_text(
            f"Varios resultados: {', '.join(m['name'] for m in matches)}. Sé más específico."
        )
        return

    emp     = matches[0]
    records = db.get_employee_records(emp["telegram_id"], limit=10)
    if not records:
        await update.message.reply_text(f"No hay registros para {emp['name']}.")
        return

    lines = [f"*Últimos registros — {emp['name']}:*"]
    for r in records:
        ent = r["entry_time"].strftime("%H:%M") if r["entry_time"] else "--:--"
        sal = r["exit_time"].strftime("%H:%M")  if r["exit_time"]  else "Sin salida"
        hs  = f"{r['total_hours']:.1f}h"        if r["total_hours"] else ""
        lines.append(f"  {r['date']}  {ent} → {sal}  {hs}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ---------- comando admin: turnos ----------

async def cmd_turno(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Uso: /turno Juan Perez 09:00 18:00"""
    if not es_admin(update.effective_user.id):
        return
    if len(context.args) < 3:
        await update.message.reply_text(
            "Uso: /turno Nombre HH:MM HH:MM\n"
            "Ejemplo: /turno Juan Perez 08:00 17:00"
        )
        return

    sal_raw = context.args[-1]
    ent_raw = context.args[-2]
    nombre  = " ".join(context.args[:-2])

    try:
        eh, em = (int(x) for x in ent_raw.split(":"))
        sh, sm = (int(x) for x in sal_raw.split(":"))
    except Exception:
        await update.message.reply_text("Formato inválido. Usá HH:MM HH:MM")
        return

    matches = db.find_employee_by_name(nombre)
    if not matches:
        await update.message.reply_text("No encontré ese empleado.")
        return
    if len(matches) > 1:
        await update.message.reply_text(
            f"Varios resultados: {', '.join(m['name'] for m in matches)}."
        )
        return

    emp = matches[0]
    db.set_shift(emp["telegram_id"], eh, em, sh, sm)
    await update.message.reply_text(
        f"Turno actualizado: {emp['name']} — {ent_raw} a {sal_raw}"
    )


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Prioridad 1: si estamos esperando el nombre, guardarlo
    if await _guardar_nombre(update, context):
        return

    # Prioridad 2: si no está registrado, pedirle el nombre
    if not db.get_employee(update.effective_user.id):
        await _pedir_nombre(update, context)
        return

    # Prioridad 3: detectar entrada/salida por palabras clave
    text = update.message.text.lower().strip()
    if any(p in text for p in PALABRAS_ENTRADA):
        await _hacer_entro(update, context)
    elif any(p in text for p in PALABRAS_SALIDA):
        await _hacer_salgo(update, context)
    else:
        await update.message.reply_text(
            "No entendí. Comandos disponibles:\n"
            "  /entro  - Registrar entrada\n"
            "  /salgo  - Registrar salida\n"
            "  /estado - Ver tu estado hoy"
        )


# ---------- comandos admin: feriados y ausencias ----------

async def cmd_feriado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Uso: /feriado DD/MM nombre del feriado"""
    if not es_admin(update.effective_user.id):
        return
    if len(context.args) < 2:
        await update.message.reply_text("Uso: /feriado DD/MM Nombre\nEjemplo: /feriado 25/12 Navidad")
        return
    try:
        day, month = context.args[0].split("/")
        d = date(ahora().year, int(month), int(day))
    except Exception:
        await update.message.reply_text("Fecha inválida. Usá el formato DD/MM")
        return
    name = " ".join(context.args[1:])
    db.add_holiday(d, name)
    await update.message.reply_text(f"Feriado agregado: {d.strftime('%d/%m/%Y')} — {name}")


async def cmd_borrarf(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Uso: /borrarf DD/MM"""
    if not es_admin(update.effective_user.id):
        return
    if not context.args:
        await update.message.reply_text("Uso: /borrarf DD/MM")
        return
    try:
        day, month = context.args[0].split("/")
        d = date(ahora().year, int(month), int(day))
    except Exception:
        await update.message.reply_text("Fecha inválida. Usá el formato DD/MM")
        return
    if db.remove_holiday(d):
        await update.message.reply_text(f"Feriado eliminado: {d.strftime('%d/%m/%Y')}")
    else:
        await update.message.reply_text(f"No había feriado registrado para {d.strftime('%d/%m/%Y')}")


async def cmd_feriados(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Lista feriados del mes actual."""
    if not es_admin(update.effective_user.id):
        return
    hoy = ahora().date()
    feriados = db.get_holidays_month(hoy.year, hoy.month)
    if not feriados:
        await update.message.reply_text(f"No hay feriados registrados para este mes.")
        return
    lines = [f"*Feriados {MESES_ES[hoy.month]} {hoy.year}:*"]
    for d_str, name in sorted(feriados.items()):
        d = date.fromisoformat(d_str)
        lines.append(f"  {d.strftime('%d/%m')} — {name}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_ausencia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Uso: /ausencia Juan Perez vacacion 25/06"""
    if not es_admin(update.effective_user.id):
        return
    if len(context.args) < 3:
        await update.message.reply_text(
            "Uso: /ausencia Nombre tipo DD/MM\n"
            "Tipos: vacacion, enfermedad, licencia, justificada, injustificada\n"
            "Ejemplo: /ausencia Juan Perez vacacion 25/06"
        )
        return

    # El tipo es la penúltima palabra, la fecha es la última
    tipo_raw = context.args[-2].lower()
    fecha_raw = context.args[-1]
    nombre = " ".join(context.args[:-2])

    if tipo_raw not in ABSENCE_TYPES:
        await update.message.reply_text(
            f"Tipo inválido: {tipo_raw}\n"
            "Tipos válidos: vacacion, enfermedad, licencia, justificada, injustificada"
        )
        return

    try:
        day, month = fecha_raw.split("/")
        d = date(ahora().year, int(month), int(day))
    except Exception:
        await update.message.reply_text("Fecha inválida. Usá el formato DD/MM")
        return

    matches = db.find_employee_by_name(nombre)
    if not matches:
        await update.message.reply_text(f"No encontré ningún empleado con ese nombre.")
        return
    if len(matches) > 1:
        names = ", ".join(m["name"] for m in matches)
        await update.message.reply_text(f"Encontré varios: {names}\nSé más específico.")
        return

    emp = matches[0]
    db.add_absence(emp["telegram_id"], d, tipo_raw)
    label = ABSENCE_TYPES[tipo_raw]
    await update.message.reply_text(
        f"Ausencia registrada:\n{emp['name']} — {label} — {d.strftime('%d/%m/%Y')}"
    )


async def cmd_ausencias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Lista ausencias del mes actual."""
    if not es_admin(update.effective_user.id):
        return
    hoy = ahora().date()
    last = calendar.monthrange(hoy.year, hoy.month)[1]
    start = date(hoy.year, hoy.month, 1)
    end = date(hoy.year, hoy.month, last)
    ausencias = db.get_absences(start, end)

    if not ausencias:
        await update.message.reply_text("No hay ausencias registradas este mes.")
        return

    lines = [f"*Ausencias {MESES_ES[hoy.month]} {hoy.year}:*"]
    for tid, dias in ausencias.items():
        emp = db.get_employee(tid)
        name = emp["name"] if emp else f"ID {tid}"
        for d_str, tipo in sorted(dias.items()):
            d = date.fromisoformat(d_str)
            lines.append(f"  {name} — {ABSENCE_TYPES.get(tipo, tipo)} — {d.strftime('%d/%m')}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ---------- notificaciones automáticas ----------

async def job_aviso_entrada(context: CallbackContext):
    """Cada 30 min entre 09:00 y 11:00 — avisa solo a quienes llevan 30+ min de retraso según su turno."""
    hoy = ahora()
    if hoy.weekday() >= 5 or db.is_holiday(hoy.date()):
        return

    sin_entrada = db.get_employees_without_entry(hoy.date())
    if not sin_entrada:
        return

    atrasados = []
    for emp in sin_entrada:
        eh, em, _, _ = db.get_shift(emp["telegram_id"])
        # Hora de entrada programada + 30 minutos de tolerancia
        debio_entrar = hoy.replace(hour=eh, minute=em, second=0, microsecond=0)
        if hoy >= debio_entrar + timedelta(minutes=30):
            atrasados.append(emp)

    if not atrasados:
        return

    nombres = "\n".join(f"  • {e['name']}" for e in atrasados)
    for admin_id in ADMIN_IDS:
        await context.bot.send_message(
            chat_id=admin_id,
            text=f"*Sin entrada registrada hoy ({hoy.strftime('%d/%m')}):*\n{nombres}",
            parse_mode="Markdown",
        )
    for emp in atrasados:
        try:
            await context.bot.send_message(
                chat_id=emp["telegram_id"],
                text="Recordatorio: todavía no registraste tu entrada de hoy.\n"
                     "Escribí /entro o \"llegué\".",
            )
        except Exception:
            pass


async def job_aviso_salida(context: CallbackContext):
    """18:30 — recuerda registrar salida a quienes tienen entrada abierta."""
    hoy = ahora().date()
    if hoy.weekday() >= 5:
        return
    if db.is_holiday(hoy):
        return

    con_entrada_abierta = db.get_employees_with_open_entry(hoy)
    if not con_entrada_abierta:
        return

    for emp in con_entrada_abierta:
        try:
            await context.bot.send_message(
                chat_id=emp["telegram_id"],
                text=f"No olvidés registrar tu salida de hoy.\nEscribí /salgo o \"me voy\".",
            )
        except Exception:
            pass

    nombres = "\n".join(f"  • {e['name']}" for e in con_entrada_abierta)
    for admin_id in ADMIN_IDS:
        await context.bot.send_message(
            chat_id=admin_id,
            text=f"*Sin salida registrada ({hoy.strftime('%d/%m')}):*\n{nombres}",
            parse_mode="Markdown",
        )


async def job_backup(context: CallbackContext):
    """23:00 — copia la base de datos a la carpeta backups/, conserva los últimos 30."""
    db_path    = os.getenv("DB_PATH", "attendance.db")
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(db_path)), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    ts   = ahora().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(backup_dir, f"attendance_{ts}.db")
    shutil.copy2(db_path, dest)
    archivos = sorted(f for f in os.listdir(backup_dir) if f.endswith(".db"))
    for viejo in archivos[:-30]:
        os.remove(os.path.join(backup_dir, viejo))


async def job_alive(context: CallbackContext):
    """08:00 — confirma al admin que el bot está activo."""
    hoy = ahora()
    for admin_id in ADMIN_IDS:
        await context.bot.send_message(
            chat_id=admin_id,
            text=f"Bot activo — {hoy.strftime('%d/%m/%Y %H:%M')}",
        )


# ---------- reporte quincena ----------

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _sheet_name(name: str) -> str:
    for ch in r"/\?*[]:":
        name = name.replace(ch, "")
    return name[:31]


# Colores de fondo por estado del día
_COLOR = {
    "weekend":  "D9D9D9",  # gris
    "holiday":  "FFE699",  # amarillo dorado
    "vacacion": "BDD7EE",  # azul claro
    "enfermedad": "C6EFCE", # verde claro
    "licencia": "FCE4D6",  # naranja claro
    "justificada": "E2EFDA",
    "injustificada": "FFC7CE", # rojo claro
    "no_exit":  "FFEB9C",  # amarillo advertencia
    "missing":  "FFC7CE",  # rojo claro
}


def _row_fill(color_key: str):
    return PatternFill("solid", fgColor=_COLOR.get(color_key, "FFFFFF"))


def _build_xlsx(records, titulo: str, start_date=None, end_date=None,
                holidays: dict = None, absences: dict = None) -> io.BytesIO:

    holidays  = holidays  or {}
    absences  = absences  or {}
    cal_mode  = start_date is not None and end_date is not None

    # Columnas: con calendario incluye Día y Estado
    if cal_mode:
        headers   = ["Fecha", "Día", "Estado", "Entrada", "Salida", "Horas Trabajadas"]
        col_entry = 4; col_exit = 5; col_hours = 6; ncols = 6
    else:
        headers   = ["Fecha", "Día", "Entrada", "Salida", "Horas Trabajadas"]
        col_entry = 3; col_exit = 4; col_hours = 5; ncols = 5

    # Estilos base
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF")
    tot_fill  = PatternFill("solid", fgColor="1F4E79")
    tot_font  = Font(bold=True, color="FFFFFF")
    name_fill = PatternFill("solid", fgColor="2E75B6")
    name_font = Font(bold=True, color="FFFFFF", size=12)

    def style_header(ws, row):
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

    def style_total(ws, row):
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = tot_fill; c.font = tot_font

    def apply_row_color(ws, row, color_key):
        fill = _row_fill(color_key)
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = fill

    wb = openpyxl.Workbook()

    # ── Hoja Resumen ─────────────────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Resumen"
    ws_res.merge_cells(f"A1:B1")
    t = ws_res.cell(row=1, column=1, value=titulo)
    t.font = Font(bold=True, size=13); t.alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Empleado", "Total Horas"], start=1):
        c = ws_res.cell(row=2, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")

    ws_res.column_dimensions["A"].width = 26
    ws_res.column_dimensions["B"].width = 16

    records_sorted = sorted(records, key=lambda r: (r["name"], r["date"]))
    grouped = {name: list(g) for name, g in groupby(records_sorted, key=lambda r: r["name"])}

    # También incluir empleados que solo tienen ausencias (sin registros)
    if cal_mode:
        all_employees = {e["name"]: e["telegram_id"] for e in db.list_employees()}
        for name in all_employees:
            if name not in grouped:
                grouped[name] = []

    summary_row  = 3
    summary_refs = {}
    detail_total = {}

    for name in sorted(grouped.keys()):
        ws_res.cell(row=summary_row, column=1, value=name)
        summary_refs[name] = summary_row
        summary_row += 1

    grand_row = summary_row
    style_total(ws_res, grand_row)
    for col in range(1, 3):
        ws_res.cell(row=grand_row, column=col).fill = tot_fill
    ws_res.cell(row=grand_row, column=1, value="TOTAL GENERAL").font = tot_font
    ws_res.cell(row=grand_row, column=1).fill = tot_fill

    # ── Leyenda de colores en Resumen ────────────────────────────────────────
    leg_row = grand_row + 2
    ws_res.cell(row=leg_row, column=1, value="Leyenda:").font = Font(bold=True)
    leyenda = [
        ("Fin de semana", "weekend"), ("Feriado", "holiday"),
        ("Vacación", "vacacion"), ("Enfermedad", "enfermedad"),
        ("Licencia", "licencia"), ("Sin salida", "no_exit"),
        ("Ausente", "missing"),
    ]
    for i, (label, key) in enumerate(leyenda):
        r = leg_row + 1 + i
        ws_res.cell(row=r, column=1, value=label).fill = _row_fill(key)
        ws_res.cell(row=r, column=1).alignment = Alignment(horizontal="left")

    # ── Una hoja por empleado ─────────────────────────────────────────────────
    for name, emp_records in sorted(grouped.items()):
        sname = _sheet_name(name)
        ws = wb.create_sheet(title=sname)

        # Índice de registros por fecha para acceso O(1)
        by_date = {r["date"]: r for r in emp_records}

        # Ausencias de este empleado
        emp_tid  = None
        emp_list = db.find_employee_by_name(name)
        if emp_list:
            emp_tid = emp_list[0]["telegram_id"]
        emp_absences = absences.get(emp_tid, {}) if emp_tid else {}

        # Fila 1: título con nombre
        ws.merge_cells(f"A1:{chr(64+ncols)}1")
        nc = ws.cell(row=1, column=1, value=name)
        nc.fill = name_fill; nc.font = name_font
        nc.alignment = Alignment(horizontal="center")

        # Fila 2: encabezados
        for col, h in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=h)
        style_header(ws, 2)

        data_start = 3
        cur_row = data_start

        if cal_mode:
            # ── Modo calendario: todos los días del período ───────────────
            from datetime import timedelta
            current_day = start_date
            while current_day <= end_date:
                d_str    = current_day.isoformat()
                weekday  = current_day.weekday()
                day_name = DIAS_ES[weekday]
                is_we    = weekday >= 5
                holiday  = holidays.get(d_str)
                absence  = emp_absences.get(d_str)
                record   = by_date.get(d_str)

                ws.cell(cur_row, 1, current_day.strftime("%d/%m/%Y"))
                ws.cell(cur_row, 2, day_name)

                if is_we:
                    ws.cell(cur_row, 3, "Fin de semana")
                    apply_row_color(ws, cur_row, "weekend")

                elif holiday:
                    ws.cell(cur_row, 3, f"Feriado: {holiday}")
                    apply_row_color(ws, cur_row, "holiday")

                elif absence:
                    label = ABSENCE_TYPES.get(absence, absence.capitalize())
                    ws.cell(cur_row, 3, label)
                    apply_row_color(ws, cur_row, absence)

                elif record:
                    if record["entry_time"] and record["exit_time"]:
                        ws.cell(cur_row, 3, "Trabajó")
                        c = ws.cell(cur_row, col_entry, record["entry_time"].replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                        c = ws.cell(cur_row, col_exit, record["exit_time"].replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                        c = ws.cell(cur_row, col_hours,
                                    f"=(E{cur_row}-D{cur_row})*24")
                        c.number_format = "0.00"
                        c.alignment = Alignment(horizontal="center")
                    else:
                        ws.cell(cur_row, 3, "Sin salida")
                        c = ws.cell(cur_row, col_entry, record["entry_time"].replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                        ws.cell(cur_row, col_hours, "Sin salida")
                        apply_row_color(ws, cur_row, "no_exit")

                elif current_day <= ahora().date():
                    ws.cell(cur_row, 3, "Ausente")
                    apply_row_color(ws, cur_row, "missing")
                # días futuros: sin color, sin estado (pendiente)

                cur_row += 1
                current_day += timedelta(days=1)

        else:
            # ── Modo simple: solo los registros existentes ────────────────
            for r in emp_records:
                d = date.fromisoformat(r["date"])
                ws.cell(cur_row, 1, r["date"])
                ws.cell(cur_row, 2, DIAS_ES[d.weekday()])
                if r["entry_time"]:
                    c = ws.cell(cur_row, col_entry, r["entry_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                if r["exit_time"]:
                    c = ws.cell(cur_row, col_exit, r["exit_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    c = ws.cell(cur_row, col_hours,
                                f"=(D{cur_row}-C{cur_row})*24")
                    c.number_format = "0.00"
                    c.alignment = Alignment(horizontal="center")
                else:
                    ws.cell(cur_row, col_hours, "Sin salida")
                    apply_row_color(ws, cur_row, "no_exit")
                cur_row += 1

        data_end  = cur_row - 1
        total_row = cur_row

        style_total(ws, total_row)
        lc = ws.cell(total_row, ncols - 1, "TOTAL")
        lc.font = tot_font; lc.fill = tot_fill
        lc.alignment = Alignment(horizontal="right")
        hcol_letter = chr(64 + col_hours)
        tc = ws.cell(total_row, col_hours,
                     f'=SUMIF({hcol_letter}{data_start}:{hcol_letter}{data_end},'
                     f'"<>Sin salida",{hcol_letter}{data_start}:{hcol_letter}{data_end})')
        tc.number_format = "0.00"; tc.font = tot_font
        tc.fill = tot_fill; tc.alignment = Alignment(horizontal="center")

        detail_total[name] = (sname, total_row, col_hours)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        if cal_mode:
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 18
        else:
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 18

    # ── Completar Resumen con referencias a hojas ─────────────────────────────
    for name, srow in summary_refs.items():
        sname, trow, hcol = detail_total[name]
        hcol_letter = chr(64 + hcol)
        c = ws_res.cell(srow, 2, f"='{sname}'!{hcol_letter}{trow}")
        c.number_format = "0.00"; c.alignment = Alignment(horizontal="center")

    if summary_refs:
        grand_refs = ",".join([f"B{r}" for r in summary_refs.values()])
        gc = ws_res.cell(grand_row, 2, f"=SUM({grand_refs})")
    else:
        gc = ws_res.cell(grand_row, 2, 0)
    gc.number_format = "0.00"; gc.font = tot_font
    gc.fill = tot_fill; gc.alignment = Alignment(horizontal="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _enviar_email(buffer: io.BytesIO, filename: str, subject: str) -> bool:
    """Envía el XLSX por Gmail. Devuelve True si tuvo éxito."""
    email_from = os.getenv("EMAIL_FROM")
    email_pass = os.getenv("EMAIL_PASSWORD")
    email_to   = os.getenv("EMAIL_TO")

    if not all([email_from, email_pass, email_to]):
        return False  # Email no configurado, se omite silenciosamente

    msg = MIMEMultipart()
    msg["From"]    = email_from
    msg["To"]      = email_to
    msg["Subject"] = subject

    msg.attach(MIMEText(
        f"Adjunto el reporte de asistencia: {filename}\n\n"
        f"Enviado automáticamente por el sistema de asistencia.",
        "plain", "utf-8"
    ))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(buffer.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(email_from, email_pass)
        server.sendmail(email_from, email_to.split(","), msg.as_string())

    return True


async def job_quincena(context: CallbackContext):
    hoy = ahora().date()
    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
    mes = MESES_ES[hoy.month]

    if hoy.day == 15:
        start = hoy.replace(day=1)
        end = hoy
        label = f"1ra quincena {mes} {hoy.year} (1-15)"
        filename = f"asistencia_{hoy.year}{hoy.month:02d}_q1.xlsx"
    elif hoy.day == ultimo_dia:
        start = hoy.replace(day=16)
        end = hoy
        label = f"2da quincena {mes} {hoy.year} (16-{ultimo_dia})"
        filename = f"asistencia_{hoy.year}{hoy.month:02d}_q2.xlsx"
    else:
        return

    records  = db.get_records_by_period(start, end)
    holidays = db.get_holidays(start, end)
    absences = db.get_absences(start, end)
    buffer   = _build_xlsx(records, label, start_date=start, end_date=end,
                           holidays=holidays, absences=absences)

    # Enviar por Telegram a todos los admins
    for admin_id in ADMIN_IDS:
        await context.bot.send_document(
            chat_id=admin_id,
            document=buffer,
            filename=filename,
            caption=f"Reporte automático - {label}\n{len(records)} registros",
        )
        buffer.seek(0)

    # Enviar por email si está configurado
    buffer.seek(0)
    enviado = _enviar_email(buffer, filename, f"Asistencia - {label}")
    if enviado:
        for admin_id in ADMIN_IDS:
            await context.bot.send_message(
                chat_id=admin_id,
                text=f"Reporte también enviado a {os.getenv('EMAIL_TO')}",
            )


# ---------- main ----------

def main():
    import asyncio
    asyncio.set_event_loop(asyncio.new_event_loop())

    if not TOKEN:
        raise ValueError("Falta la variable de entorno TELEGRAM_TOKEN")
    if not ADMIN_IDS:
        print("ADVERTENCIA: ADMIN_IDS no configurado. Nadie podrá descargar reportes.")

    app = Application.builder().token(TOKEN).build()

    app.add_handler(CommandHandler("start",        cmd_start))
    app.add_handler(CommandHandler("entro",        cmd_entro))
    app.add_handler(CommandHandler("salgo",        cmd_salgo))
    app.add_handler(CommandHandler("estado",       cmd_estado))
    app.add_handler(CommandHandler("reporte",      cmd_reporte))
    app.add_handler(CommandHandler("empleados",    cmd_empleados))
    app.add_handler(CommandHandler("feriado",      cmd_feriado))
    app.add_handler(CommandHandler("borrarf",      cmd_borrarf))
    app.add_handler(CommandHandler("feriados",     cmd_feriados))
    app.add_handler(CommandHandler("ausencia",     cmd_ausencia))
    app.add_handler(CommandHandler("ausencias",    cmd_ausencias))
    app.add_handler(CommandHandler("corregir",     cmd_corregir))
    app.add_handler(CommandHandler("borrar",       cmd_borrar_fichaje))
    app.add_handler(CommandHandler("ver",          cmd_ver_fichajes))
    app.add_handler(CommandHandler("turno",        cmd_turno))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND,     handle_text))
    app.add_handler(MessageHandler(filters.LOCATION,                    handle_location))

    jq = app.job_queue
    # Reporte quincena: día 15 y último del mes a las 18:00
    jq.run_daily(job_quincena,    time=dt_time(hour=18, minute=0,  tzinfo=TIMEZONE))
    # Avisos de entrada: cada hora entre 09:00 y 12:00, respeta turnos individuales
    for h in (9, 10, 11, 12):
        jq.run_daily(job_aviso_entrada, time=dt_time(hour=h, minute=0, tzinfo=TIMEZONE))
    # Aviso de salida pendiente a las 18:30
    jq.run_daily(job_aviso_salida, time=dt_time(hour=18, minute=30, tzinfo=TIMEZONE))
    # Backup diario a las 23:00
    jq.run_daily(job_backup,       time=dt_time(hour=23, minute=0,  tzinfo=TIMEZONE))
    # Confirmación de vida a las 08:00
    jq.run_daily(job_alive,        time=dt_time(hour=8,  minute=0,  tzinfo=TIMEZONE))

    print("Bot iniciado. Esperando mensajes...")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
