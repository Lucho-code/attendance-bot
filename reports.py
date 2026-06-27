"""
GeneraciÃ³n de reportes XLSX y envÃ­o de email.
Compartido entre bot.py y admin_panel.py.
"""
import io
import os
import smtplib
from datetime import date, datetime, timedelta
from itertools import groupby
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pytz

TIMEZONE = pytz.timezone("America/Argentina/Buenos_Aires")

DIAS_ES = ["Lunes", "Martes", "MiÃ©rcoles", "Jueves", "Viernes", "SÃ¡bado", "Domingo"]

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

ABSENCE_TYPES = {
    "vacacion": "VacaciÃ³n", "vacaciÃ³n": "VacaciÃ³n",
    "enfermedad": "Enfermedad",
    "licencia": "Licencia",
    "justificada": "Justificada",
    "injustificada": "Injustificada",
}

_COLOR = {
    "weekend":       "D9D9D9",
    "holiday":       "FFE699",
    "vacacion":      "BDD7EE",
    "enfermedad":    "C6EFCE",
    "licencia":      "FCE4D6",
    "justificada":   "E2EFDA",
    "injustificada": "FFC7CE",
    "no_exit":       "FFEB9C",
    "missing":       "FFC7CE",
    "overtime":      "FFD966",
}


def _row_fill(key: str) -> PatternFill:
    return PatternFill("solid", fgColor=_COLOR.get(key, "FFFFFF"))


def _sheet_name(name: str) -> str:
    for ch in r"/\?*[]:":
        name = name.replace(ch, "")
    return name[:31]


def _scheduled_hours(shift: tuple) -> float:
    eh, em, sh, sm = shift
    return (sh * 60 + sm - eh * 60 - em) / 60


import math as _math
def _r15(h: float) -> float:
    """Ceil al cuarto de hora: siempre redondea hacia arriba al 0.25 mas cercano.
    Nunca pierde minutos de horas extra. Ej: 0.25->0.25, 0.49->0.50, 0.75->0.75, 4.95->5.00"""
    return _math.ceil(h * 4) / 4


GRACE_MINUTES = int(os.getenv("GRACE_MINUTES", "15"))


def calcular_horas(entry_dt: datetime, exit_dt: datetime,
                   weekday: int, is_holiday: bool) -> tuple:
    """
    Devuelve (horas_normales, horas_extra_50, horas_extra_100).

    Reglas de horas:
      Lun-Vie 07:00-16:00          â†’ normales
      Lun-Vie fuera de ese rango   â†’ extra 50%
      SÃ¡bado  07:00-13:00          â†’ extra 50%
      SÃ¡bado  13:00+ / antes 07:00 â†’ extra 100%
      Domingo / Feriado             â†’ extra 100%

    Tolerancia (GRACE_MINUTES, default 15 min):
      Si la entrada cae dentro de los primeros GRACE minutos de una zona,
      se ajusta al inicio de esa zona (no se cobra como extra).
      Si la salida cae dentro de los Ãºltimos GRACE minutos de una zona,
      se ajusta al fin de esa zona (se cuenta como jornada completa).
    """
    entry = entry_dt.replace(tzinfo=None)
    exit_ = exit_dt.replace(tzinfo=None)
    grace = timedelta(minutes=GRACE_MINUTES)
    base  = entry.replace(hour=0, minute=0, second=0, microsecond=0)

    def _boundary(h: int, m: int) -> datetime:
        return base.replace(hour=h, minute=m, second=0, microsecond=0)

    def _snap(dt: datetime, h: int, m: int) -> datetime:
        """Ajusta dt al lÃ­mite h:m si cae dentro del margen de gracia."""
        b = _boundary(h, m)
        return b if abs(dt - b) <= grace else dt

    if is_holiday or weekday == 6:
        total = (exit_ - entry).total_seconds() / 3600
        return (0.0, 0.0, _r15(total))

    if weekday == 5:    # SÃ¡bado: zona 07:00 y frontera 13:00
        entry = _snap(entry, 7,  0)
        exit_ = _snap(exit_,  7,  0)
        exit_ = _snap(exit_,  13, 0)
    else:               # Lun-Vie: zona 07:00-16:00
        entry = _snap(entry, 7,  0)
        exit_ = _snap(exit_,  16, 0)

    if exit_ <= entry:
        return (0.0, 0.0, 0.0)

    def _overlap(zone_h0: int, zone_m0: int, zone_h1: int, zone_m1: int) -> float:
        zs = _boundary(zone_h0, zone_m0)
        ze = _boundary(zone_h1, zone_m1)
        s  = max(entry, zs)
        e  = min(exit_,  ze)
        return max(0.0, (e - s).total_seconds() / 3600)

    if weekday == 5:
        extra_50  = _overlap(7,  0, 13, 0)
        extra_100 = _overlap(0,  0,  7, 0) + _overlap(13, 0, 23, 59)
        return (0.0, _r15(extra_50), _r15(extra_100))

    normal   = _overlap(7,  0, 16, 0)
    extra_50 = _overlap(0,  0,  7, 0) + _overlap(16, 0, 23, 59)
    return (_r15(normal), _r15(extra_50), 0.0)
    """
    Genera el XLSX de asistencia con desglose de horas normales,
    extra 50% y extra 100% segÃºn convenio.
    """
    holidays = holidays or {}
    absences = absences or {}
    cal_mode = start_date is not None and end_date is not None
    today    = datetime.now(TIMEZONE).date()

    if cal_mode:
        headers   = ["Fecha", "DÃ­a", "Estado", "Entrada", "Salida",
                     "Hs. Normales", "Hs. Extra 50%", "Hs. Extra 100%"]
        col_entry = 4; col_exit = 5
        col_norm  = 6; col_50 = 7; col_100 = 8; ncols = 8
    else:
        headers   = ["Fecha", "DÃ­a", "Entrada", "Salida",
                     "Hs. Normales", "Hs. Extra 50%", "Hs. Extra 100%"]
        col_entry = 3; col_exit = 4
        col_norm  = 5; col_50 = 6; col_100 = 7; ncols = 7

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF")
    tot_fill  = PatternFill("solid", fgColor="1F4E79")
    tot_font  = Font(bold=True, color="FFFFFF")
    name_fill = PatternFill("solid", fgColor="2E75B6")
    name_font = Font(bold=True, color="FFFFFF", size=12)

    def style_hdr(ws, row):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

    def style_tot(ws, row):
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = tot_fill
            ws.cell(row=row, column=c).font = tot_font

    def color_row(ws, row, key):
        fill = _row_fill(key)
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = fill

    wb = openpyxl.Workbook()

    # â”€â”€ Hoja Resumen â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws_res = wb.active
    ws_res.title = "Resumen"
    ws_res.merge_cells("A1:C1")
    t = ws_res.cell(1, 1, titulo)
    t.font = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Empleado", "Hs. Normales", "Hs. Extra 50%", "Hs. Extra 100%"],
                            start=1):
        c = ws_res.cell(2, col, h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    ws_res.column_dimensions["A"].width = 26
    ws_res.column_dimensions["B"].width = 14
    ws_res.column_dimensions["C"].width = 14
    ws_res.column_dimensions["D"].width = 16

    records_sorted = sorted(records, key=lambda r: (r["name"], r["date"]))
    grouped = {n: list(g)
               for n, g in groupby(records_sorted, key=lambda r: r["name"])}

    if cal_mode:
        for e in db.list_employees():
            if e["name"] not in grouped:
                grouped[e["name"]] = []

    summary_row  = 3
    summary_refs = {}
    detail_info  = {}   # name -> (sheet_name, total_row, extra_row)

    for name in sorted(grouped):
        ws_res.cell(summary_row, 1, name)
        summary_refs[name] = summary_row
        summary_row += 1

    grand_row = summary_row
    style_tot(ws_res, grand_row)
    for c in range(1, 5):
        ws_res.cell(grand_row, c).fill = tot_fill
    ws_res.cell(grand_row, 1, "TOTAL GENERAL").font = tot_font
    ws_res.cell(grand_row, 1).fill = tot_fill

    # Leyenda
    leg_row = grand_row + 2
    ws_res.cell(leg_row, 1, "Leyenda:").font = Font(bold=True)
    leyenda = [
        ("Fin de semana", "weekend"), ("Feriado", "holiday"),
        ("VacaciÃ³n", "vacacion"),     ("Enfermedad", "enfermedad"),
        ("Licencia", "licencia"),     ("Sin salida", "no_exit"),
        ("Ausente", "missing"),       ("Horas extra", "overtime"),
    ]
    for i, (label, key) in enumerate(leyenda):
        r = leg_row + 1 + i
        ws_res.cell(r, 1, label).fill = _row_fill(key)

    # â”€â”€ Una hoja por empleado â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    for name, emp_records in sorted(grouped.items()):
        sname = _sheet_name(name)
        ws    = wb.create_sheet(title=sname)

        by_date  = {r["date"]: r for r in emp_records}
        emp_list = db.find_employee_by_name(name)
        emp_tid  = emp_list[0]["telegram_id"] if emp_list else None
        emp_abs  = absences.get(emp_tid, {}) if emp_tid else {}

        ws.merge_cells(f"A1:{chr(64+ncols)}1")
        nc = ws.cell(1, 1, name)
        nc.fill = name_fill; nc.font = name_font
        nc.alignment = Alignment(horizontal="center")

        for col, h in enumerate(headers, start=1):
            ws.cell(2, col, h)
        style_hdr(ws, 2)

        data_start = 3
        cur_row    = data_start

        def _write_hours(row, norm, e50, e100):
            has_extra = e50 > 0 or e100 > 0
            for col, val in ((col_norm, norm), (col_50, e50), (col_100, e100)):
                if val > 0:
                    c = ws.cell(row, col, val)
                    c.number_format = "0.00"
                    c.alignment = Alignment(horizontal="center")
            if has_extra:
                color_row(ws, row, "overtime")

        if cal_mode:
            current_day = start_date
            while current_day <= end_date:
                d_str    = current_day.isoformat()
                weekday  = current_day.weekday()
                is_we    = weekday >= 5
                holiday  = holidays.get(d_str)
                absence  = emp_abs.get(d_str)
                record   = by_date.get(d_str)
                is_hday  = bool(holiday)

                ws.cell(cur_row, 1, current_day.strftime("%d/%m/%Y"))
                ws.cell(cur_row, 2, DIAS_ES[weekday])

                if is_we and not is_hday:
                    ws.cell(cur_row, 3, "Fin de semana")
                    color_row(ws, cur_row, "weekend")

                elif is_hday:
                    ws.cell(cur_row, 3, f"Feriado: {holiday}")
                    color_row(ws, cur_row, "holiday")
                    # Si hay registro en feriado â†’ calcular horas (todas al 100%)
                    if record and record.get("entry_time") and record.get("exit_time"):
                        c = ws.cell(cur_row, col_entry,
                                    record["entry_time"].replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                        c = ws.cell(cur_row, col_exit,
                                    record["exit_time"].replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                        n, e50, e100 = calcular_horas(
                            record["entry_time"], record["exit_time"], weekday, True)
                        _write_hours(cur_row, n, e50, e100)

                elif absence:
                    ws.cell(cur_row, 3, ABSENCE_TYPES.get(absence, absence.capitalize()))
                    color_row(ws, cur_row, absence)

                elif record and record.get("entry_time") and record.get("exit_time"):
                    # Sumar horas de todas las sesiones del dÃ­a
                    sessions  = record.get("sessions", [record])
                    n = e50 = e100 = 0.0
                    for s in sessions:
                        if s.get("entry_time") and s.get("exit_time"):
                            sn, se50, se100 = calcular_horas(
                                s["entry_time"], s["exit_time"], weekday, False)
                            n += sn; e50 += se50; e100 += se100
                    n_ses = record.get("session_count", 1)
                    label = f"TrabajÃ³ ({n_ses} salidas)" if n_ses > 1 else "TrabajÃ³"
                    ws.cell(cur_row, 3, label)
                    c = ws.cell(cur_row, col_entry,
                                record["entry_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    c = ws.cell(cur_row, col_exit,
                                record["exit_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    _write_hours(cur_row, round(n,2), round(e50,2), round(e100,2))

                elif record and record.get("entry_time"):
                    ws.cell(cur_row, 3, "Sin salida")
                    c = ws.cell(cur_row, col_entry,
                                record["entry_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    color_row(ws, cur_row, "no_exit")

                elif current_day <= today and not is_we:
                    ws.cell(cur_row, 3, "Ausente")
                    color_row(ws, cur_row, "missing")

                cur_row     += 1
                current_day += timedelta(days=1)

        else:
            for r in emp_records:
                d = date.fromisoformat(r["date"])
                ws.cell(cur_row, 1, r["date"])
                ws.cell(cur_row, 2, DIAS_ES[d.weekday()])
                if r.get("entry_time"):
                    c = ws.cell(cur_row, col_entry,
                                r["entry_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                if r.get("exit_time"):
                    c = ws.cell(cur_row, col_exit,
                                r["exit_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    is_hday = bool(holidays.get(r["date"]))
                    n, e50, e100 = calcular_horas(
                        r["entry_time"], r["exit_time"], d.weekday(), is_hday)
                    _write_hours(cur_row, n, e50, e100)
                else:
                    color_row(ws, cur_row, "no_exit")
                cur_row += 1

        data_end  = cur_row - 1
        total_row = cur_row
        style_tot(ws, total_row)

        lbl = ws.cell(total_row, col_norm - 1, "TOTAL")
        lbl.font = tot_font; lbl.fill = tot_fill
        lbl.alignment = Alignment(horizontal="right")

        for col in (col_norm, col_50, col_100):
            lc = chr(64 + col)
            tc = ws.cell(total_row, col,
                         f"=SUM({lc}{data_start}:{lc}{data_end})")
            tc.number_format = "0.00"; tc.font = tot_font
            tc.fill = tot_fill; tc.alignment = Alignment(horizontal="center")

        detail_info[name] = (sname, total_row)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        if cal_mode:
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 13
            ws.column_dimensions["G"].width = 12
        else:
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 13
            ws.column_dimensions["F"].width = 12

    # â”€â”€ Completar Resumen â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    for name, srow in summary_refs.items():
        if name not in detail_info:
            continue
        sname, trow = detail_info[name]
        for res_col, emp_col in ((2, col_norm), (3, col_50), (4, col_100)):
            lc = chr(64 + emp_col)
            c  = ws_res.cell(srow, res_col, f"='{sname}'!{lc}{trow}")
            c.number_format = "0.00"; c.alignment = Alignment(horizontal="center")

    if summary_refs:
        for res_col, col_letter in ((2, "B"), (3, "C"), (4, "D")):
            refs = ",".join(f"{col_letter}{r}" for r in summary_refs.values())
            gc   = ws_res.cell(grand_row, res_col, f"=SUM({refs})")
            gc.number_format = "0.00"; gc.font = tot_font
            gc.fill = tot_fill; gc.alignment = Alignment(horizontal="center")
    else:
        for c in (2, 3, 4):
            ws_res.cell(grand_row, c, 0).font = tot_font
            ws_res.cell(grand_row, c).fill = tot_fill

    # â”€â”€ Hoja Obras (si hay datos) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if start_date and end_date:
        try:
            obras_rows = db.conn.execute("""
                SELECT ob.name obra_name, e.name emp_name,
                       os.date, os.entry_time, os.exit_time, os.total_hours
                FROM obra_sessions os
                JOIN employees e  ON os.telegram_id = e.telegram_id
                JOIN obras ob ON os.obra_id = ob.id
                WHERE os.date BETWEEN ? AND ?
                ORDER BY ob.name, os.date, e.name
            """, (start_date.isoformat(), end_date.isoformat())).fetchall()

            if obras_rows:
                ws_o = wb.create_sheet(title="Obras")
                for col, h in enumerate(["Obra", "Empleado", "Fecha",
                                         "Entrada", "Salida", "Horas"], start=1):
                    c = ws_o.cell(1, col, h)
                    c.fill = hdr_fill; c.font = hdr_font
                    c.alignment = Alignment(horizontal="center")

                for i, r in enumerate(obras_rows, start=2):
                    ws_o.cell(i, 1, r["obra_name"])
                    ws_o.cell(i, 2, r["emp_name"])
                    ws_o.cell(i, 3, r["date"])
                    ent = r["entry_time"]
                    sal = r["exit_time"]
                    if ent:
                        dt = datetime.fromisoformat(ent)
                        c  = ws_o.cell(i, 4, (TIMEZONE.localize(dt) if dt.tzinfo is None
                                               else dt.astimezone(TIMEZONE)).replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                    if sal:
                        dt = datetime.fromisoformat(sal)
                        c  = ws_o.cell(i, 5, (TIMEZONE.localize(dt) if dt.tzinfo is None
                                               else dt.astimezone(TIMEZONE)).replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                    if r["total_hours"]:
                        ws_o.cell(i, 6, round(r["total_hours"], 2)).number_format = "0.00"

                for ltr, w in zip("ABCDEF", [28, 22, 12, 10, 10, 10]):
                    ws_o.column_dimensions[ltr].width = w

                # Totales por obra
                from collections import defaultdict
                tot_obra = defaultdict(float)
                for r in obras_rows:
                    if r["total_hours"]:
                        tot_obra[r["obra_name"]] += r["total_hours"]

                tot_row = len(obras_rows) + 3
                ws_o.cell(tot_row, 1, "TOTAL POR OBRA").font = Font(bold=True)
                for j, (nombre, horas) in enumerate(sorted(tot_obra.items()), start=tot_row + 1):
                    ws_o.cell(j, 1, nombre)
                    ws_o.cell(j, 6, round(horas, 2)).number_format = "0.00"
        except Exception:
            pass  # Si algo falla, el resto del reporte sigue igual

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_xlsx(db, records, titulo: str,
               start_date: date = None, end_date: date = None,
               holidays: dict = None, absences: dict = None) -> io.BytesIO:
    """
    Genera el XLSX de asistencia con desglose de horas normales,
    extra 50% y extra 100% según convenio.
    """
    holidays = holidays or {}
    absences = absences or {}
    cal_mode = start_date is not None and end_date is not None
    today    = datetime.now(TIMEZONE).date()

    if cal_mode:
        headers   = ["Fecha", "Día", "Estado", "Entrada", "Salida",
                     "Hs. Normales", "Hs. Extra 50%", "Hs. Extra 100%"]
        col_entry = 4; col_exit = 5
        col_norm  = 6; col_50 = 7; col_100 = 8; ncols = 8
    else:
        headers   = ["Fecha", "Día", "Entrada", "Salida",
                     "Hs. Normales", "Hs. Extra 50%", "Hs. Extra 100%"]
        col_entry = 3; col_exit = 4
        col_norm  = 5; col_50 = 6; col_100 = 7; ncols = 7

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF")
    tot_fill  = PatternFill("solid", fgColor="1F4E79")
    tot_font  = Font(bold=True, color="FFFFFF")
    name_fill = PatternFill("solid", fgColor="2E75B6")
    name_font = Font(bold=True, color="FFFFFF", size=12)

    def style_hdr(ws, row):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

    def style_tot(ws, row):
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = tot_fill
            ws.cell(row=row, column=c).font = tot_font

    def color_row(ws, row, key):
        fill = _row_fill(key)
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = fill

    wb = openpyxl.Workbook()

    # ── Hoja Resumen ──────────────────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Resumen"
    ws_res.merge_cells("A1:C1")
    t = ws_res.cell(1, 1, titulo)
    t.font = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Empleado", "Hs. Normales", "Hs. Extra 50%", "Hs. Extra 100%"],
                            start=1):
        c = ws_res.cell(2, col, h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    ws_res.column_dimensions["A"].width = 26
    ws_res.column_dimensions["B"].width = 14
    ws_res.column_dimensions["C"].width = 14
    ws_res.column_dimensions["D"].width = 16

    records_sorted = sorted(records, key=lambda r: (r["name"], r["date"]))
    grouped = {n: list(g)
               for n, g in groupby(records_sorted, key=lambda r: r["name"])}

    if cal_mode:
        for e in db.list_employees():
            if e["name"] not in grouped:
                grouped[e["name"]] = []

    summary_row  = 3
    summary_refs = {}
    detail_info  = {}   # name -> (sheet_name, total_row, extra_row)

    for name in sorted(grouped):
        ws_res.cell(summary_row, 1, name)
        summary_refs[name] = summary_row
        summary_row += 1

    grand_row = summary_row
    style_tot(ws_res, grand_row)
    for c in range(1, 5):
        ws_res.cell(grand_row, c).fill = tot_fill
    ws_res.cell(grand_row, 1, "TOTAL GENERAL").font = tot_font
    ws_res.cell(grand_row, 1).fill = tot_fill

    # Leyenda
    leg_row = grand_row + 2
    ws_res.cell(leg_row, 1, "Leyenda:").font = Font(bold=True)
    leyenda = [
        ("Fin de semana", "weekend"), ("Feriado", "holiday"),
        ("Vacación", "vacacion"),     ("Enfermedad", "enfermedad"),
        ("Licencia", "licencia"),     ("Sin salida", "no_exit"),
        ("Ausente", "missing"),       ("Horas extra", "overtime"),
    ]
    for i, (label, key) in enumerate(leyenda):
        r = leg_row + 1 + i
        ws_res.cell(r, 1, label).fill = _row_fill(key)

    # ── Una hoja por empleado ─────────────────────────────────────────────────
    for name, emp_records in sorted(grouped.items()):
        sname = _sheet_name(name)
        ws    = wb.create_sheet(title=sname)

        by_date  = {r["date"]: r for r in emp_records}
        emp_list = db.find_employee_by_name(name)
        emp_tid  = emp_list[0]["telegram_id"] if emp_list else None
        emp_abs  = absences.get(emp_tid, {}) if emp_tid else {}

        ws.merge_cells(f"A1:{chr(64+ncols)}1")
        nc = ws.cell(1, 1, name)
        nc.fill = name_fill; nc.font = name_font
        nc.alignment = Alignment(horizontal="center")

        for col, h in enumerate(headers, start=1):
            ws.cell(2, col, h)
        style_hdr(ws, 2)

        data_start = 3
        cur_row    = data_start

        def _write_hours(row, norm, e50, e100):
            has_extra = e50 > 0 or e100 > 0
            for col, val in ((col_norm, norm), (col_50, e50), (col_100, e100)):
                if val > 0:
                    c = ws.cell(row, col, val)
                    c.number_format = "0.00"
                    c.alignment = Alignment(horizontal="center")
            if has_extra:
                color_row(ws, row, "overtime")

        if cal_mode:
            current_day = start_date
            while current_day <= end_date:
                d_str    = current_day.isoformat()
                weekday  = current_day.weekday()
                is_we    = weekday >= 5
                holiday  = holidays.get(d_str)
                absence  = emp_abs.get(d_str)
                record   = by_date.get(d_str)
                is_hday  = bool(holiday)

                ws.cell(cur_row, 1, current_day.strftime("%d/%m/%Y"))
                ws.cell(cur_row, 2, DIAS_ES[weekday])

                if is_we and not is_hday:
                    ws.cell(cur_row, 3, "Fin de semana")
                    color_row(ws, cur_row, "weekend")

                elif is_hday:
                    ws.cell(cur_row, 3, f"Feriado: {holiday}")
                    color_row(ws, cur_row, "holiday")
                    # Si hay registro en feriado → calcular horas (todas al 100%)
                    if record and record.get("entry_time") and record.get("exit_time"):
                        c = ws.cell(cur_row, col_entry,
                                    record["entry_time"].replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                        c = ws.cell(cur_row, col_exit,
                                    record["exit_time"].replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                        n, e50, e100 = calcular_horas(
                            record["entry_time"], record["exit_time"], weekday, True)
                        _write_hours(cur_row, n, e50, e100)

                elif absence:
                    ws.cell(cur_row, 3, ABSENCE_TYPES.get(absence, absence.capitalize()))
                    color_row(ws, cur_row, absence)

                elif record and record.get("entry_time") and record.get("exit_time"):
                    # Sumar horas de todas las sesiones del día
                    sessions  = record.get("sessions", [record])
                    n = e50 = e100 = 0.0
                    for s in sessions:
                        if s.get("entry_time") and s.get("exit_time"):
                            sn, se50, se100 = calcular_horas(
                                s["entry_time"], s["exit_time"], weekday, False)
                            n += sn; e50 += se50; e100 += se100
                    n_ses = record.get("session_count", 1)
                    label = f"Trabajó ({n_ses} salidas)" if n_ses > 1 else "Trabajó"
                    ws.cell(cur_row, 3, label)
                    c = ws.cell(cur_row, col_entry,
                                record["entry_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    c = ws.cell(cur_row, col_exit,
                                record["exit_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    _write_hours(cur_row, round(n,2), round(e50,2), round(e100,2))

                elif record and record.get("entry_time"):
                    ws.cell(cur_row, 3, "Sin salida")
                    c = ws.cell(cur_row, col_entry,
                                record["entry_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    color_row(ws, cur_row, "no_exit")

                elif current_day <= today and not is_we:
                    ws.cell(cur_row, 3, "Ausente")
                    color_row(ws, cur_row, "missing")

                cur_row     += 1
                current_day += timedelta(days=1)

        else:
            for r in emp_records:
                d = date.fromisoformat(r["date"])
                ws.cell(cur_row, 1, r["date"])
                ws.cell(cur_row, 2, DIAS_ES[d.weekday()])
                if r.get("entry_time"):
                    c = ws.cell(cur_row, col_entry,
                                r["entry_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                if r.get("exit_time"):
                    c = ws.cell(cur_row, col_exit,
                                r["exit_time"].replace(tzinfo=None).time())
                    c.number_format = "HH:MM"
                    is_hday = bool(holidays.get(r["date"]))
                    n, e50, e100 = calcular_horas(
                        r["entry_time"], r["exit_time"], d.weekday(), is_hday)
                    _write_hours(cur_row, n, e50, e100)
                else:
                    color_row(ws, cur_row, "no_exit")
                cur_row += 1

        data_end  = cur_row - 1
        total_row = cur_row
        style_tot(ws, total_row)

        lbl = ws.cell(total_row, col_norm - 1, "TOTAL")
        lbl.font = tot_font; lbl.fill = tot_fill
        lbl.alignment = Alignment(horizontal="right")

        for col in (col_norm, col_50, col_100):
            lc = chr(64 + col)
            tc = ws.cell(total_row, col,
                         f"=SUM({lc}{data_start}:{lc}{data_end})")
            tc.number_format = "0.00"; tc.font = tot_font
            tc.fill = tot_fill; tc.alignment = Alignment(horizontal="center")

        detail_info[name] = (sname, total_row)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        if cal_mode:
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 13
            ws.column_dimensions["G"].width = 12
        else:
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 13
            ws.column_dimensions["F"].width = 12

    # ── Completar Resumen ─────────────────────────────────────────────────────
    for name, srow in summary_refs.items():
        if name not in detail_info:
            continue
        sname, trow = detail_info[name]
        for res_col, emp_col in ((2, col_norm), (3, col_50), (4, col_100)):
            lc = chr(64 + emp_col)
            c  = ws_res.cell(srow, res_col, f"='{sname}'!{lc}{trow}")
            c.number_format = "0.00"; c.alignment = Alignment(horizontal="center")

    if summary_refs:
        for res_col, col_letter in ((2, "B"), (3, "C"), (4, "D")):
            refs = ",".join(f"{col_letter}{r}" for r in summary_refs.values())
            gc   = ws_res.cell(grand_row, res_col, f"=SUM({refs})")
            gc.number_format = "0.00"; gc.font = tot_font
            gc.fill = tot_fill; gc.alignment = Alignment(horizontal="center")
    else:
        for c in (2, 3, 4):
            ws_res.cell(grand_row, c, 0).font = tot_font
            ws_res.cell(grand_row, c).fill = tot_fill

    # ── Hoja Obras (si hay datos) ─────────────────────────────────────────────
    if start_date and end_date:
        try:
            obras_rows = db.conn.execute("""
                SELECT ob.name obra_name, e.name emp_name,
                       os.date, os.entry_time, os.exit_time, os.total_hours
                FROM obra_sessions os
                JOIN employees e  ON os.telegram_id = e.telegram_id
                JOIN obras ob ON os.obra_id = ob.id
                WHERE os.date BETWEEN ? AND ?
                ORDER BY ob.name, os.date, e.name
            """, (start_date.isoformat(), end_date.isoformat())).fetchall()

            if obras_rows:
                ws_o = wb.create_sheet(title="Obras")
                for col, h in enumerate(["Obra", "Empleado", "Fecha",
                                         "Entrada", "Salida", "Horas"], start=1):
                    c = ws_o.cell(1, col, h)
                    c.fill = hdr_fill; c.font = hdr_font
                    c.alignment = Alignment(horizontal="center")

                for i, r in enumerate(obras_rows, start=2):
                    ws_o.cell(i, 1, r["obra_name"])
                    ws_o.cell(i, 2, r["emp_name"])
                    ws_o.cell(i, 3, r["date"])
                    ent = r["entry_time"]
                    sal = r["exit_time"]
                    if ent:
                        dt = datetime.fromisoformat(ent)
                        c  = ws_o.cell(i, 4, (TIMEZONE.localize(dt) if dt.tzinfo is None
                                               else dt.astimezone(TIMEZONE)).replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                    if sal:
                        dt = datetime.fromisoformat(sal)
                        c  = ws_o.cell(i, 5, (TIMEZONE.localize(dt) if dt.tzinfo is None
                                               else dt.astimezone(TIMEZONE)).replace(tzinfo=None).time())
                        c.number_format = "HH:MM"
                    if r["total_hours"]:
                        ws_o.cell(i, 6, round(r["total_hours"], 2)).number_format = "0.00"

                for ltr, w in zip("ABCDEF", [28, 22, 12, 10, 10, 10]):
                    ws_o.column_dimensions[ltr].width = w

                # Totales por obra
                from collections import defaultdict
                tot_obra = defaultdict(float)
                for r in obras_rows:
                    if r["total_hours"]:
                        tot_obra[r["obra_name"]] += r["total_hours"]

                tot_row = len(obras_rows) + 3
                ws_o.cell(tot_row, 1, "TOTAL POR OBRA").font = Font(bold=True)
                for j, (nombre, horas) in enumerate(sorted(tot_obra.items()), start=tot_row + 1):
                    ws_o.cell(j, 1, nombre)
                    ws_o.cell(j, 6, round(horas, 2)).number_format = "0.00"
        except Exception:
            pass  # Si algo falla, el resto del reporte sigue igual

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def send_email(buf: io.BytesIO, filename: str, subject: str) -> bool:
    email_from = os.getenv("EMAIL_FROM")
    email_pass = os.getenv("EMAIL_PASSWORD")
    email_to   = os.getenv("EMAIL_TO")
    if not all([email_from, email_pass, email_to]):
        return False
    msg = MIMEMultipart()
    msg["From"] = email_from
    msg["To"]   = email_to
    msg["Subject"] = subject
    msg.attach(MIMEText(
        f"Adjunto el reporte de asistencia: {filename}\n\n"
        "Enviado automáticamente por el sistema.",
        "plain", "utf-8",
    ))
    part = MIMEBase("application", "octet-stream")
    part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(email_from, email_pass)
        s.sendmail(email_from, email_to.split(","), msg.as_string())
    return True
